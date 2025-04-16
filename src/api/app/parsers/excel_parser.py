"""
Excel file parser for the API service.

This module provides a parser for Excel files (xlsx, xls, csv) that extracts
structured data from them. It handles the initial parsing and validation of Excel
files before delegating complex processing to the processor service.
"""
import os
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Union
import io
import tempfile
import csv
import chardet

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter

from app.parsers.base_parser import BaseParser
from app.utils.logger import get_logger

# Initialize logger
logger = get_logger("excel_parser")


class ExcelParser(BaseParser):
    """
    Parser for Excel files (xlsx, xls, csv).
    
    This parser extracts data from Excel files and converts it to a structured
    format that can be processed by the extraction service.
    """
    
    SUPPORTED_FILE_TYPES = ["xlsx", "xls", "csv"]
    DEFAULT_CSV_ENCODING = 'utf-8'
    DEFAULT_CHUNK_SIZE = 10000  # For large file processing
    
    async def parse(self, file_path: Union[str, Path], file_type: Optional[str] = None) -> Dict[str, Any]:
        """
        Parse an Excel file and return structured data.
        
        Args:
            file_path: Path to the Excel file
            file_type: Optional file type override
            
        Returns:
            Dict containing the parsed Excel data with sheet information
            
        Raises:
            ValueError: If the file cannot be parsed as an Excel file
            FileNotFoundError: If the file does not exist
        """
        file_path = Path(file_path)
        
        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")
        
        if not await self.validate(file_path, file_type):
            raise ValueError(f"Invalid file type: {file_path}")
        
        file_extension = file_type or file_path.suffix.lower().lstrip(".")
        
        try:
            if file_extension == "csv":
                # Detect encoding for CSV files
                encoding = await self._detect_csv_encoding(file_path)
                delimiter = await self._detect_csv_delimiter(file_path, encoding)
                
                # Check file size for chunking
                file_size = os.path.getsize(file_path)
                if file_size > 10 * 1024 * 1024:  # 10MB
                    # Process large CSV in chunks
                    df = await self._process_large_csv(file_path, encoding, delimiter)
                else:
                    df = pd.read_csv(file_path, encoding=encoding, delimiter=delimiter)
                
                sheets = {"Sheet1": await self._dataframe_to_dict(df)}
                metadata = {
                    "encoding": encoding,
                    "delimiter": delimiter,
                    "has_header": self._detect_header(df),
                }
            else:  # xlsx or xls
                # Use context manager for file handling
                xlsx = pd.ExcelFile(file_path)
                sheets = {}
                metadata = await self._extract_excel_metadata(file_path)
                
                for sheet_name in xlsx.sheet_names:
                    # Check if sheet is empty
                    try:
                        df = pd.read_excel(xlsx, sheet_name=sheet_name)
                        if df.shape[0] > 0 and df.shape[1] > 0:
                            # Process sheet data
                            sheets[sheet_name] = await self._dataframe_to_dict(df, sheet_name)
                    except Exception as sheet_error:
                        logger.warning(f"Error reading sheet {sheet_name}", error=str(sheet_error), file=file_path.name)
                        # Continue with other sheets if one fails
                        continue
            
            return {
                "file_name": file_path.name,
                "file_type": file_extension,
                "sheets": sheets,
                "sheet_names": list(sheets.keys()),
                "metadata": metadata
            }
        
        except Exception as e:
            logger.error(f"Error parsing Excel file", file=file_path.name, error=str(e))
            raise ValueError(f"Error parsing Excel file: {str(e)}")
    
    async def validate(self, file_path: Union[str, Path], file_type: Optional[str] = None) -> bool:
        """
        Validate that the file is a valid Excel file.
        
        Args:
            file_path: Path to the Excel file
            file_type: Optional file type override
            
        Returns:
            True if the file is a valid Excel file, False otherwise
        """
        file_path = Path(file_path)
        
        if not file_path.exists():
            return False
        
        file_extension = file_type or file_path.suffix.lower().lstrip(".")
        
        if file_extension not in self.SUPPORTED_FILE_TYPES:
            return False
        
        try:
            # Basic validation to ensure file can be opened
            if file_extension == "csv":
                # Try to detect encoding first
                encoding = await self._detect_csv_encoding(file_path)
                pd.read_csv(file_path, nrows=5, encoding=encoding)
            else:  # xlsx or xls
                # Check if file is corrupt
                if file_extension == "xlsx":
                    try:
                        load_workbook(file_path, read_only=True)
                    except Exception as openpyxl_error:
                        logger.warning(f"OpenPyXL validation failed", file=file_path.name, error=str(openpyxl_error))
                        # Fall back to pandas validation
                
                pd.read_excel(file_path, nrows=5)
            return True
        except Exception as e:
            logger.error(f"Error validating Excel file", file=file_path.name, error=str(e))
            return False
    
    def get_supported_file_types(self) -> List[str]:
        """
        Get the list of file types supported by this parser.
        
        Returns:
            List of supported file extensions (without the dot)
        """
        return self.SUPPORTED_FILE_TYPES
    
    async def _dataframe_to_dict(self, df: pd.DataFrame, sheet_name: str = None) -> Dict[str, Any]:
        """
        Convert a pandas DataFrame to a dictionary suitable for processing.
        
        Args:
            df: Pandas DataFrame to convert
            sheet_name: Optional sheet name for additional context
            
        Returns:
            Dict representation of the DataFrame
        """
        # Detect and handle merged cells (will be reflected in metadata)
        merged_cells_info = []
        if sheet_name and hasattr(df, '_mgr') and hasattr(df._mgr, 'blocks'):
            # This is a best-effort approach - real merged cell data requires openpyxl
            # We'll add actual merged cell handling in the metadata extraction
            pass
        
        # Improve handling of various data types
        # Convert DataFrame to dictionary with handling for various data types
        # Replace NaN, NA, None values with None for JSON compatibility
        df_cleaned = df.copy()
        
        # Handle different data types appropriately
        for col in df_cleaned.columns:
            if pd.api.types.is_datetime64_any_dtype(df_cleaned[col]):
                # Convert datetime to ISO format strings
                df_cleaned[col] = df_cleaned[col].dt.strftime('%Y-%m-%dT%H:%M:%S')
            elif pd.api.types.is_numeric_dtype(df_cleaned[col]):
                # Keep numeric columns as is
                pass
            else:
                # Convert other types to string, but handle None/NaN
                df_cleaned[col] = df_cleaned[col].astype(str).replace('nan', None).replace('None', None)
        
        # Convert to records
        records = df_cleaned.replace({pd.NA: None, np.nan: None}).to_dict(orient="records")
        
        # Get column types and names with more detailed type information
        columns = []
        for col in df.columns:
            col_type = str(df[col].dtype)
            sample_values = df[col].dropna().head(3).tolist()
            
            columns.append({
                "name": str(col),
                "type": col_type,
                "sample_values": sample_values if sample_values else [],
                "unique_count": df[col].nunique() if not df[col].empty else 0,
                "null_count": df[col].isna().sum()
            })
        
        # Return enhanced structured data
        return {
            "data": records,
            "columns": columns,
            "shape": {
                "rows": df.shape[0],
                "columns": df.shape[1],
            },
            "merged_cells": merged_cells_info,
            "statistical_summary": {
                "numeric_columns": {
                    col: {
                        "min": float(df[col].min()) if pd.api.types.is_numeric_dtype(df[col]) and not df[col].empty else None,
                        "max": float(df[col].max()) if pd.api.types.is_numeric_dtype(df[col]) and not df[col].empty else None,
                        "mean": float(df[col].mean()) if pd.api.types.is_numeric_dtype(df[col]) and not df[col].empty else None,
                    }
                    for col in df.columns if pd.api.types.is_numeric_dtype(df[col])
                }
            }
        }
    
    async def _detect_csv_encoding(self, file_path: Union[str, Path]) -> str:
        """
        Detect the encoding of a CSV file.
        
        Args:
            file_path: Path to the CSV file
            
        Returns:
            Detected encoding
        """
        # Read a small portion of the file to detect encoding
        with open(file_path, 'rb') as f:
            raw_data = f.read(min(1024 * 1024, os.path.getsize(file_path)))  # Read up to 1MB
        
        result = chardet.detect(raw_data)
        encoding = result['encoding'] if result['encoding'] else self.DEFAULT_CSV_ENCODING
        confidence = result['confidence']
        
        logger.debug(f"Detected CSV encoding", encoding=encoding, confidence=confidence, file=file_path.name)
        return encoding
    
    async def _detect_csv_delimiter(self, file_path: Union[str, Path], encoding: str) -> str:
        """
        Detect the delimiter of a CSV file.
        
        Args:
            file_path: Path to the CSV file
            encoding: File encoding
            
        Returns:
            Detected delimiter
        """
        # Common delimiters to check
        delimiters = [',', ';', '\t', '|']
        
        # Read a small portion of the file
        with open(file_path, 'r', encoding=encoding, errors='replace') as f:
            sample = f.read(4096)
        
        # Count occurrences of each delimiter in the sample
        counts = {delimiter: sample.count(delimiter) for delimiter in delimiters}
        
        # Return the delimiter with the highest count
        delimiter = max(counts.items(), key=lambda x: x[1])[0]
        logger.debug(f"Detected CSV delimiter", delimiter=delimiter, file=file_path.name)
        return delimiter
    
    async def _process_large_csv(self, file_path: Union[str, Path], encoding: str, delimiter: str) -> pd.DataFrame:
        """
        Process a large CSV file in chunks.
        
        Args:
            file_path: Path to the CSV file
            encoding: File encoding
            delimiter: CSV delimiter
            
        Returns:
            DataFrame with sample data from the file
        """
        # For large files, just read a sample for preview
        chunks = []
        
        # Read first chunk and last chunk to provide a sample
        first_chunk = pd.read_csv(
            file_path, 
            encoding=encoding, 
            delimiter=delimiter,
            nrows=self.DEFAULT_CHUNK_SIZE // 2
        )
        chunks.append(first_chunk)
        
        # If the file has more rows, read a sample from the end
        file_size = os.path.getsize(file_path)
        if file_size > 20 * 1024 * 1024:  # Very large file
            # Read the number of rows in the file
            with open(file_path, 'r', encoding=encoding, errors='replace') as f:
                num_rows = sum(1 for _ in f) - 1  # Subtract header row
                
            if num_rows > self.DEFAULT_CHUNK_SIZE:
                # Read a sample from the middle/end
                skiprows = max(1, num_rows - (self.DEFAULT_CHUNK_SIZE // 2))
                last_chunk = pd.read_csv(
                    file_path,
                    encoding=encoding,
                    delimiter=delimiter,
                    skiprows=range(1, skiprows)
                )
                chunks.append(last_chunk)
        
        # Combine the chunks
        return pd.concat(chunks) if len(chunks) > 1 else chunks[0]
    
    async def _extract_excel_metadata(self, file_path: Union[str, Path]) -> Dict[str, Any]:
        """
        Extract metadata from an Excel file.
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            Dict containing metadata
        """
        metadata = {}
        
        # Only extract detailed metadata for xlsx files using openpyxl
        if Path(file_path).suffix.lower() == '.xlsx':
            try:
                # Use read_only mode for performance with large files
                wb = load_workbook(filename=file_path, read_only=False, data_only=True)
                
                # Document properties
                metadata["properties"] = {
                    "author": wb.properties.creator,
                    "created": wb.properties.created.isoformat() if wb.properties.created else None,
                    "modified": wb.properties.modified.isoformat() if wb.properties.modified else None,
                    "title": wb.properties.title,
                    "subject": wb.properties.subject,
                }
                
                # Sheet information
                metadata["sheets"] = {}
                
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    
                    # Collect merged cells
                    merged_cells = []
                    for merged_cell_range in sheet.merged_cells.ranges:
                        merged_cells.append(str(merged_cell_range))
                    
                    # Get sheet dimensions
                    sheet_metadata = {
                        "dimension": sheet.calculate_dimension(),
                        "merged_cells": merged_cells,
                        "has_formula": False,  # Will set to True if formulas found
                    }
                    
                    # Sample checks for formulas (can be slow for large sheets)
                    formula_count = 0
                    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=min(100, sheet.max_row))):
                        for cell in row:
                            if cell.data_type == 'f':  # Formula cell
                                formula_count += 1
                                sheet_metadata["has_formula"] = True
                    
                    sheet_metadata["formula_sample_count"] = formula_count
                    metadata["sheets"][sheet_name] = sheet_metadata
                
                wb.close()
            except Exception as e:
                logger.warning(f"Error extracting Excel metadata", file=file_path.name, error=str(e))
                # Continue without detailed metadata
        
        return metadata
    
    def _detect_header(self, df: pd.DataFrame) -> bool:
        """
        Attempt to detect if the first row is a header.
        
        Args:
            df: DataFrame to analyze
            
        Returns:
            True if the first row is likely a header, False otherwise
        """
        if df.shape[0] <= 1:
            return True  # Default to True for very small files
            
        # Check if first row has different data types than the rest
        if all(isinstance(col, str) for col in df.columns):
            # For numeric data, headers are often strings while data is numeric
            numeric_cols = 0
            for col in df.columns:
                if pd.api.types.is_numeric_dtype(df[col][1:]):
                    if not pd.api.types.is_numeric_dtype(pd.Series([df.iloc[0][col]])):
                        numeric_cols += 1
            
            # If most columns follow this pattern, likely has header
            return numeric_cols > len(df.columns) / 2
            
        return True  # Default to assuming there is a header 