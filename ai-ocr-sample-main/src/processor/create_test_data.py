#!/usr/bin/env python
"""
Generate sample Excel files for testing the AI-OCR system.
This script creates various types of test Excel files to validate the extraction pipeline.
"""
import argparse
import os
import random
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def parse_args():
    """Parse command line arguments."""
    parser = argparse.ArgumentParser(description="Generate test Excel files for AI-OCR system")
    parser.add_argument("--output-dir", type=str, default="test_data",
                        help="Directory to store generated test files")
    parser.add_argument("--num-files", type=int, default=1,
                        help="Number of files to generate for each type")
    return parser.parse_args()


def create_output_dir(output_dir):
    """Create output directory if it doesn't exist."""
    output_path = Path(output_dir)
    output_path.mkdir(exist_ok=True, parents=True)
    return output_path


def generate_invoice(file_path, index=1):
    """Generate a sample invoice Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"
    
    # Add company header
    ws.merge_cells('A1:E1')
    ws['A1'] = "ACME CORPORATION"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Add invoice title
    ws.merge_cells('A3:E3')
    ws['A3'] = "INVOICE"
    ws['A3'].font = Font(bold=True, size=12)
    ws['A3'].alignment = Alignment(horizontal='center')
    
    # Add invoice metadata
    today = datetime.now()
    due_date = today + timedelta(days=30)
    invoice_number = f"INV-{today.year}-{index:03d}"
    
    ws['A5'] = "Invoice Number:"
    ws['B5'] = invoice_number
    ws['A6'] = "Date:"
    ws['B6'] = today.strftime("%Y-%m-%d")
    ws['A7'] = "Due Date:"
    ws['B7'] = due_date.strftime("%Y-%m-%d")
    
    # Add customer information
    ws['D5'] = "Customer:"
    ws['E5'] = f"Customer {index}"
    ws['D6'] = "Account #:"
    ws['E6'] = f"CUST-{random.randint(10000, 99999)}"
    ws['D7'] = "Reference:"
    ws['E7'] = f"REF-{random.randint(10000, 99999)}"
    
    # Add line items header
    headers = ["Description", "Quantity", "Unit Price", "Tax", "Amount"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=9, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    # Add line items
    products = [
        "Product A", 
        "Product B", 
        "Service C", 
        "Maintenance D",
        "Support E"
    ]
    
    num_items = random.randint(3, 5)
    total_amount = 0
    
    for i in range(num_items):
        row = 10 + i
        product = products[i % len(products)]
        qty = random.randint(1, 10)
        unit_price = round(random.uniform(100, 1000), 2)
        tax_rate = 0.1  # 10% tax
        tax = round(qty * unit_price * tax_rate, 2)
        amount = round(qty * unit_price + tax, 2)
        total_amount += amount
        
        ws.cell(row=row, column=1, value=f"{product} {i+1}")
        ws.cell(row=row, column=2, value=qty)
        ws.cell(row=row, column=3, value=unit_price)
        ws.cell(row=row, column=4, value=tax)
        ws.cell(row=row, column=5, value=amount)
    
    # Add totals
    subtotal_row = 10 + num_items + 1
    tax_row = subtotal_row + 1
    total_row = tax_row + 1
    
    subtotal = round(total_amount / 1.1, 2)  # Remove tax to get subtotal
    tax_total = round(total_amount - subtotal, 2)
    
    ws.cell(row=subtotal_row, column=4, value="Subtotal:")
    ws.cell(row=subtotal_row, column=5, value=subtotal)
    
    ws.cell(row=tax_row, column=4, value="Tax:")
    ws.cell(row=tax_row, column=5, value=tax_total)
    
    ws.cell(row=total_row, column=4, value="Total:")
    ws.cell(row=total_row, column=5, value=total_amount)
    ws.cell(row=total_row, column=4).font = Font(bold=True)
    ws.cell(row=total_row, column=5).font = Font(bold=True)
    
    # Add payment information
    payment_row = total_row + 2
    ws.merge_cells(f'A{payment_row}:E{payment_row}')
    ws[f'A{payment_row}'] = "Payment Information"
    ws[f'A{payment_row}'].font = Font(bold=True)
    
    ws[f'A{payment_row+1}'] = "Bank Name:"
    ws[f'B{payment_row+1}'] = "ACME Bank"
    ws[f'A{payment_row+2}'] = "Account Number:"
    ws[f'B{payment_row+2}'] = f"ACCT-{random.randint(100000, 999999)}"
    ws[f'A{payment_row+3}'] = "Routing Number:"
    ws[f'B{payment_row+3}'] = f"RTG-{random.randint(100000, 999999)}"
    
    # Add vendor information
    vendor_row = payment_row + 5
    ws.merge_cells(f'A{vendor_row}:E{vendor_row}')
    ws[f'A{vendor_row}'] = "Vendor Information"
    ws[f'A{vendor_row}'].font = Font(bold=True)
    
    ws[f'A{vendor_row+1}'] = "Name:"
    ws[f'B{vendor_row+1}'] = "ACME Corporation"
    ws[f'A{vendor_row+2}'] = "Address:"
    ws[f'B{vendor_row+2}'] = "123 ACME Street, ACME City, AC 12345"
    ws[f'A{vendor_row+3}'] = "Tax ID:"
    ws[f'B{vendor_row+3}'] = f"TAX-{random.randint(100000, 999999)}"
    
    # Adjust column widths
    for col in range(1, 6):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = 15
    
    # Save the workbook
    wb.save(file_path)
    return invoice_number, total_amount


def generate_sales_report(file_path, index=1):
    """Generate a sample sales report Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Report"
    
    # Add report title
    report_date = datetime.now() - timedelta(days=random.randint(1, 30))
    ws.merge_cells('A1:E1')
    ws['A1'] = f"Monthly Sales Report - {report_date.strftime('%B %Y')}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Add report metadata
    ws['A3'] = "Report Date:"
    ws['B3'] = report_date.strftime("%Y-%m-%d")
    ws['A4'] = "Report ID:"
    ws['B4'] = f"REP-{report_date.strftime('%Y%m')}-{index:03d}"
    ws['A5'] = "Author:"
    ws['B5'] = f"User {random.randint(100, 999)}"
    
    # Add executive summary
    ws.merge_cells('A7:E7')
    ws['A7'] = "Executive Summary"
    ws['A7'].font = Font(bold=True)
    
    ws.merge_cells('A8:E10')
    ws['A8'] = (
        f"This report summarizes sales performance for {report_date.strftime('%B %Y')}. "
        f"Overall sales reached {random.randint(100000, 999999)} units, representing a "
        f"{random.randint(5, 20)}% increase compared to the previous month."
    )
    ws['A8'].alignment = Alignment(wrap_text=True)
    
    # Add sales by product category
    ws.merge_cells('A12:E12')
    ws['A12'] = "Sales by Product Category"
    ws['A12'].font = Font(bold=True)
    
    # Add category headers
    headers = ["Category", "Q1", "Q2", "Q3", "Q4", "Total"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=13, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    # Add category data
    categories = ["Electronics", "Furniture", "Office Supplies", "Services", "Software"]
    total_sales = 0
    
    for i, category in enumerate(categories):
        row = 14 + i
        q1 = random.randint(10000, 50000)
        q2 = random.randint(10000, 50000)
        q3 = random.randint(10000, 50000)
        q4 = random.randint(10000, 50000)
        total = q1 + q2 + q3 + q4
        total_sales += total
        
        ws.cell(row=row, column=1, value=category)
        ws.cell(row=row, column=2, value=q1)
        ws.cell(row=row, column=3, value=q2)
        ws.cell(row=row, column=4, value=q3)
        ws.cell(row=row, column=5, value=q4)
        ws.cell(row=row, column=6, value=total)
    
    # Add total row
    total_row = 14 + len(categories)
    ws.cell(row=total_row, column=1, value="Total")
    total_q1 = sum(ws.cell(row=14+i, column=2).value for i in range(len(categories)))
    total_q2 = sum(ws.cell(row=14+i, column=3).value for i in range(len(categories)))
    total_q3 = sum(ws.cell(row=14+i, column=4).value for i in range(len(categories)))
    total_q4 = sum(ws.cell(row=14+i, column=5).value for i in range(len(categories)))
    
    ws.cell(row=total_row, column=2, value=total_q1)
    ws.cell(row=total_row, column=3, value=total_q2)
    ws.cell(row=total_row, column=4, value=total_q3)
    ws.cell(row=total_row, column=5, value=total_q4)
    ws.cell(row=total_row, column=6, value=total_sales)
    
    for col in range(1, 7):
        ws.cell(row=total_row, column=col).font = Font(bold=True)
    
    # Add regional sales data in a new sheet
    ws2 = wb.create_sheet(title="Regional Data")
    
    # Add sheet title
    ws2.merge_cells('A1:D1')
    ws2['A1'] = "Regional Sales Data"
    ws2['A1'].font = Font(bold=True, size=12)
    ws2['A1'].alignment = Alignment(horizontal='center')
    
    # Add region headers
    region_headers = ["Region", "Sales", "Market Share", "YoY Growth"]
    for col, header in enumerate(region_headers, start=1):
        cell = ws2.cell(row=3, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    # Add region data
    regions = ["North", "South", "East", "West", "International"]
    total_regional_sales = 0
    
    for i, region in enumerate(regions):
        row = 4 + i
        sales = random.randint(50000, 200000)
        market_share = f"{random.randint(10, 35)}%"
        yoy_growth = f"{random.randint(-5, 25)}%"
        total_regional_sales += sales
        
        ws2.cell(row=row, column=1, value=region)
        ws2.cell(row=row, column=2, value=sales)
        ws2.cell(row=row, column=3, value=market_share)
        ws2.cell(row=row, column=4, value=yoy_growth)
    
    # Add third sheet with monthly trends
    ws3 = wb.create_sheet(title="Monthly Trends")
    
    # Add sheet title
    ws3.merge_cells('A1:D1')
    ws3['A1'] = "Monthly Sales Trends"
    ws3['A1'].font = Font(bold=True, size=12)
    ws3['A1'].alignment = Alignment(horizontal='center')
    
    # Create monthly data
    months = ["January", "February", "March", "April", "May", "June", 
              "July", "August", "September", "October", "November", "December"]
    
    # Add headers
    ws3.cell(row=3, column=1, value="Month")
    ws3.cell(row=3, column=2, value="Sales")
    ws3.cell(row=3, column=2).font = Font(bold=True)
    
    # Add monthly data
    for i, month in enumerate(months):
        row = 4 + i
        sales = random.randint(80000, 150000)
        
        ws3.cell(row=row, column=1, value=month)
        ws3.cell(row=row, column=2, value=sales)
    
    # Adjust column widths
    for sheet in wb.worksheets:
        for col in range(1, 7):
            column_letter = get_column_letter(col)
            sheet.column_dimensions[column_letter].width = 15
    
    # Save the workbook
    wb.save(file_path)
    return report_date.strftime('%B %Y'), total_sales


def generate_product_catalog(file_path, index=1):
    """Generate a sample product catalog Excel file."""
    # Create DataFrame for products
    num_products = random.randint(20, 50)
    
    data = {
        "Product ID": [f"PRD-{random.randint(1000, 9999)}" for _ in range(num_products)],
        "Product Name": [f"Product {i+1}" for i in range(num_products)],
        "Category": [random.choice(["Electronics", "Furniture", "Office Supplies", "Software", "Hardware"]) for _ in range(num_products)],
        "Price": [round(random.uniform(10, 1000), 2) for _ in range(num_products)],
        "Stock": [random.randint(0, 1000) for _ in range(num_products)],
        "Description": [f"Description for product {i+1}" for i in range(num_products)]
    }
    
    df = pd.DataFrame(data)
    
    # Create DataFrame for categories
    categories = ["Electronics", "Furniture", "Office Supplies", "Software", "Hardware"]
    category_data = {
        "Category": categories,
        "Total Products": [len(df[df["Category"] == cat]) for cat in categories],
        "Average Price": [round(df[df["Category"] == cat]["Price"].mean(), 2) for cat in categories],
        "Total Stock": [df[df["Category"] == cat]["Stock"].sum() for cat in categories],
    }
    
    df_categories = pd.DataFrame(category_data)
    
    # Create Excel file with multiple sheets
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        # Add products sheet
        df.to_excel(writer, sheet_name='Products', index=False)
        
        # Add categories sheet
        df_categories.to_excel(writer, sheet_name='Categories', index=False)
        
        # Add top products sheet (sample of top 10 products by price)
        top_products = df.sort_values(by='Price', ascending=False).head(10)
        top_products.to_excel(writer, sheet_name='Top Products', index=False)
    
    return num_products, len(categories)


def main():
    """Main function to generate test files."""
    args = parse_args()
    output_dir = create_output_dir(args.output_dir)
    
    print(f"Generating {args.num_files} files of each type in {output_dir}")
    
    # Generate invoices
    for i in range(1, args.num_files + 1):
        invoice_path = output_dir / f"invoice_{i}.xlsx"
        invoice_number, total_amount = generate_invoice(invoice_path, i)
        print(f"Generated invoice: {invoice_path} - Invoice #{invoice_number}, Total: ${total_amount:.2f}")
    
    # Generate sales reports
    for i in range(1, args.num_files + 1):
        report_path = output_dir / f"sales_report_{i}.xlsx"
        report_month, total_sales = generate_sales_report(report_path, i)
        print(f"Generated sales report: {report_path} - Period: {report_month}, Total Sales: ${total_sales:.2f}")
    
    # Generate product catalogs
    for i in range(1, args.num_files + 1):
        catalog_path = output_dir / f"product_catalog_{i}.xlsx"
        num_products, num_categories = generate_product_catalog(catalog_path, i)
        print(f"Generated product catalog: {catalog_path} - Products: {num_products}, Categories: {num_categories}")


if __name__ == "__main__":
    main() 