"""
Tests for the AI extraction functionality.
"""
import asyncio
import json
import os
import uuid
from pathlib import Path
from typing import Dict, Any

import pandas as pd
import pytest
from openpyxl import Workbook

from app.parsers.excel_parser import ExcelParser
from app.extractors.langchain_extractor import LangChainExtractor
from app.validators.schema_validator import SchemaValidator


class TestExtractionPipeline:
    """Test the extraction pipeline."""

    @pytest.fixture
    def config(self):
        """Get test configuration."""
        return {
            "model_name": os.environ.get("MODEL_NAME", "gpt-4o"),
            "model_temperature": float(os.environ.get("MODEL_TEMPERATURE", "0.1")),
            "model_api_key": os.environ.get("MODEL_API_KEY", ""),
            "test_data_dir": os.environ.get("TEST_DATA_DIR", "test_data"),
        }

    @pytest.fixture
    def excel_parser(self):
        """Create Excel parser instance."""
        return ExcelParser()

    @pytest.fixture
    def langchain_extractor(self, config):
        """Create LangChain extractor instance."""
        if not config.get("model_api_key"):
            pytest.skip("MODEL_API_KEY not set, skipping AI extraction tests")
        
        return LangChainExtractor(
            model_name=config["model_name"],
            model_temperature=config["model_temperature"],
            api_key=config["model_api_key"],
        )

    @pytest.fixture
    def schema_validator(self):
        """Create schema validator instance."""
        return SchemaValidator()

    @pytest.fixture
    def sample_invoice_path(self, tmp_path):
        """Create a sample invoice Excel file for testing."""
        file_path = tmp_path / "sample_invoice.xlsx"
        
        # Create workbook with sample invoice data
        wb = Workbook()
        ws = wb.active
        ws.title = "Invoice"
        
        # Add headers
        headers = ["Invoice Number", "Date", "Due Date", "Customer", "Total Amount"]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)
        
        # Add sample data
        data = ["INV-2025-001", "2025-03-26", "2025-04-25", "ACME Corporation", 12345.67]
        for col, value in enumerate(data, start=1):
            ws.cell(row=2, column=col, value=value)
        
        # Add line items section
        ws.cell(row=4, column=1, value="Line Items")
        
        item_headers = ["Description", "Quantity", "Unit Price", "Amount"]
        for col, header in enumerate(item_headers, start=1):
            ws.cell(row=5, column=col, value=header)
        
        # Sample line items
        line_items = [
            ["Product A", 10, 100.00, 1000.00],
            ["Service B", 5, 200.00, 1000.00],
            ["Maintenance", 1, 10345.67, 10345.67]
        ]
        
        for row_idx, item in enumerate(line_items, start=6):
            for col_idx, value in enumerate(item, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Add vendor information
        ws.cell(row=10, column=1, value="Vendor Information")
        ws.cell(row=11, column=1, value="Name:")
        ws.cell(row=11, column=2, value="Example Vendor Inc.")
        ws.cell(row=12, column=1, value="Address:")
        ws.cell(row=12, column=2, value="123 Vendor Street, Vendor City")
        ws.cell(row=13, column=1, value="Tax ID:")
        ws.cell(row=13, column=2, value="TAX-12345-678")
        
        # Save the workbook
        wb.save(file_path)
        return file_path

    @pytest.fixture
    def sample_report_path(self, tmp_path):
        """Create a sample report Excel file for testing."""
        file_path = tmp_path / "sample_report.xlsx"
        
        # Create workbook with sample report data
        wb = Workbook()
        ws = wb.active
        ws.title = "Monthly Report"
        
        # Add title and metadata
        ws.cell(row=1, column=1, value="Monthly Sales Report")
        ws.cell(row=2, column=1, value="Date:")
        ws.cell(row=2, column=2, value="2025-03-01")
        ws.cell(row=3, column=1, value="Author:")
        ws.cell(row=3, column=2, value="John Doe")
        
        # Add sales data
        ws.cell(row=5, column=1, value="Sales by Product Category")
        
        # Headers
        categories_headers = ["Category", "Q1", "Q2", "Q3", "Q4", "Total"]
        for col, header in enumerate(categories_headers, start=1):
            ws.cell(row=6, column=col, value=header)
        
        # Data
        categories_data = [
            ["Electronics", 10000, 12000, 15000, 20000, 57000],
            ["Furniture", 8000, 7500, 8200, 9000, 32700],
            ["Office Supplies", 5000, 5200, 5100, 5300, 20600],
            ["Services", 12000, 13000, 14000, 15000, 54000]
        ]
        
        for row_idx, item in enumerate(categories_data, start=7):
            for col_idx, value in enumerate(item, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Add monthly trend
        ws.cell(row=12, column=1, value="Monthly Sales Trend")
        
        # Create a new sheet for regional data
        ws2 = wb.create_sheet(title="Regional Data")
        
        # Add region headers
        ws2.cell(row=1, column=1, value="Regional Sales Data")
        region_headers = ["Region", "Sales", "Market Share", "YoY Growth"]
        for col, header in enumerate(region_headers, start=1):
            ws2.cell(row=2, column=col, value=header)
        
        # Add region data
        region_data = [
            ["North", 25000, "35%", "12%"],
            ["South", 18000, "25%", "8%"],
            ["East", 15000, "21%", "5%"],
            ["West", 14000, "19%", "7%"]
        ]
        
        for row_idx, item in enumerate(region_data, start=3):
            for col_idx, value in enumerate(item, start=1):
                ws2.cell(row=row_idx, column=col_idx, value=value)
        
        # Save the workbook
        wb.save(file_path)
        return file_path

    @pytest.fixture
    def empty_file_path(self, tmp_path):
        """Create an empty Excel file for testing."""
        file_path = tmp_path / "empty.xlsx"
        wb = Workbook()
        wb.save(file_path)
        return file_path

    @pytest.mark.asyncio
    async def test_invoice_extraction(self, excel_parser, langchain_extractor, schema_validator, sample_invoice_path):
        """Test the extraction of an invoice document."""
        # Parse the Excel file
        parsed_data = await excel_parser.parse(str(sample_invoice_path), "xlsx")
        
        # Verify parsing results
        assert isinstance(parsed_data, dict)
        assert "sheets" in parsed_data
        assert "Invoice" in parsed_data["sheets"]
        
        # Extract data using LangChain
        extracted_data, confidence_score = await langchain_extractor.extract(parsed_data)
        
        # Verify extraction results
        assert isinstance(extracted_data, dict)
        assert isinstance(confidence_score, float)
        assert confidence_score > 0.5  # Assuming reasonable confidence
        
        # Verify key fields
        assert "invoice_number" in extracted_data
        assert extracted_data["invoice_number"] == "INV-2025-001"
        assert "date" in extracted_data
        assert "total_amount" in extracted_data
        assert abs(float(extracted_data["total_amount"]) - 12345.67) < 0.01
        
        # Check vendor information
        assert "vendor" in extracted_data
        assert "name" in extracted_data["vendor"]
        assert extracted_data["vendor"]["name"] == "Example Vendor Inc."
        
        # Check line items
        assert "line_items" in extracted_data
        assert isinstance(extracted_data["line_items"], list)
        assert len(extracted_data["line_items"]) >= 3
        
        # Validate the data
        validation_result = await schema_validator.validate(extracted_data)
        
        # Verify validation results
        assert isinstance(validation_result, dict)
        assert "valid" in validation_result
        assert validation_result["valid"] is True
        assert validation_result["schema_type"] == "invoice"

    @pytest.mark.asyncio
    async def test_report_extraction(self, excel_parser, langchain_extractor, schema_validator, sample_report_path):
        """Test the extraction of a report document."""
        # Parse the Excel file
        parsed_data = await excel_parser.parse(str(sample_report_path), "xlsx")
        
        # Verify parsing results
        assert isinstance(parsed_data, dict)
        assert "sheets" in parsed_data
        assert "Monthly Report" in parsed_data["sheets"]
        assert "Regional Data" in parsed_data["sheets"]
        
        # Extract data using LangChain
        extracted_data, confidence_score = await langchain_extractor.extract(parsed_data)
        
        # Verify extraction results
        assert isinstance(extracted_data, dict)
        assert isinstance(confidence_score, float)
        assert confidence_score > 0.5  # Assuming reasonable confidence
        
        # Verify key fields
        assert "title" in extracted_data
        assert "Monthly Sales Report" in extracted_data["title"]
        assert "date" in extracted_data
        
        # Check if data extraction worked
        assert "data" in extracted_data
        assert "categories" in extracted_data["data"] or "sales" in extracted_data["data"]
        
        # Validate the data
        validation_result = await schema_validator.validate(extracted_data)
        
        # Verify validation results
        assert isinstance(validation_result, dict)
        assert "valid" in validation_result
        assert validation_result["schema_type"] == "report"

    @pytest.mark.asyncio
    async def test_empty_file_handling(self, excel_parser, langchain_extractor, schema_validator, empty_file_path):
        """Test handling of an empty file."""
        # Parse the Excel file
        parsed_data = await excel_parser.parse(str(empty_file_path), "xlsx")
        
        # Verify parsing results for empty file
        assert isinstance(parsed_data, dict)
        assert "sheets" in parsed_data
        assert "Sheet" in parsed_data["sheets"]
        assert parsed_data["sheets"]["Sheet"]["empty"] is True
        
        # Extract data using LangChain (should handle empty data gracefully)
        extracted_data, confidence_score = await langchain_extractor.extract(parsed_data)
        
        # Verify extraction results - might be minimal but shouldn't crash
        assert isinstance(extracted_data, dict)
        assert isinstance(confidence_score, float)
        assert confidence_score < 0.5  # Confidence should be low for empty files
        
        # Validate the data - should identify as invalid or minimal structure
        validation_result = await schema_validator.validate(extracted_data)
        
        # Verify validation results
        assert isinstance(validation_result, dict)
        assert "valid" in validation_result
        # Either invalid or minimal form structure
        if validation_result["valid"]:
            assert validation_result["schema_type"] == "form"

    @pytest.mark.asyncio
    async def test_error_handling(self, excel_parser, schema_validator):
        """Test error handling in the extraction pipeline."""
        # Create an invalid extractor (missing API key)
        invalid_extractor = LangChainExtractor(
            model_name="non-existent-model",
            model_temperature=0.1,
            api_key="invalid-key"
        )
        
        # Create test data
        test_data = {
            "metadata": {"filename": "test.xlsx", "sheet_names": ["Sheet1"]},
            "sheets": {
                "Sheet1": {
                    "empty": False,
                    "rows": 1,
                    "columns": 1,
                    "data": [{"Column1": "Value1"}]
                }
            }
        }
        
        # Test error handling in extraction
        with pytest.raises(Exception):
            await invalid_extractor.extract(test_data)
        
        # Test validation with invalid data
        invalid_data = {"invalid_field": "test"}
        validation_result = await schema_validator.validate(invalid_data)
        
        # Should not crash, but identify as invalid
        assert isinstance(validation_result, dict)
        assert "valid" in validation_result
        assert validation_result["valid"] is False 